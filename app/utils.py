import openpyxl
from django.core.mail import send_mail
from openpyxl.styles import Font
from datetime import datetime
from .models import Update
from core.settings import MEDIA_ROOT, DEFAULT_FROM_EMAIL


def export_to_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = '№'
    ws['B1'] = 'Name'
    ws['C1'] = 'Link'
    ws['D1'] = 'Current Location'
    ws['E1'] = 'Destination'
    ws['F1'] = 'Planned Time'
    ws['G1'] = 'ETA'
    ws['H1'] = 'ETA Status'
    ws['I1'] = 'Status'
    ws['J1'] = 'Fuel'
    ws['K1'] = 'Updated At'
    ws['L1'] = 'Live Share'

    bold_font = Font(bold=True)
    ws['A1'].font = bold_font
    ws['B1'].font = bold_font
    ws['C1'].font = bold_font
    ws['D1'].font = bold_font
    ws['E1'].font = bold_font
    ws['F1'].font = bold_font
    ws['G1'].font = bold_font
    ws['H1'].font = bold_font
    ws['I1'].font = bold_font
    ws['J1'].font = bold_font
    ws['K1'].font = bold_font
    ws['L1'].font = bold_font

    updates = [update for update in Update.objects.all()]
    rows = 2
    i = 1
    for update in updates:
        ws[f'A{rows}'] = i
        ws[f'B{rows}'] = update.name
        ws[f'C{rows}'] = update.link
        ws[f'D{rows}'] = update.current_location
        ws[f'E{rows}'] = update.destination
        ws[f'F{rows}'] = update.planned_time
        ws[f'G{rows}'] = update.eta
        ws[f'H{rows}'] = update.eta_status
        ws[f'I{rows}'] = update.status
        ws[f'J{rows}'] = update.fuel
        ws[f'K{rows}'] = str(update.updated_at)
        ws[f'L{rows}'] = update.live_share
        rows += 1
        i += 1
    excel_filename = f'table_{datetime.now().timestamp()}.xlsx'
    excel_path = f'{MEDIA_ROOT}/tables/{excel_filename}'
    wb.save(excel_path)
    return excel_path, excel_filename


def send_mail_to_broker(mail, subject, context):
    send_mail(subject, context, DEFAULT_FROM_EMAIL, [mail])
